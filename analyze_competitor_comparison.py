#!/usr/bin/env python3
"""
競合比較分析スクリプト
競合.xlsxと担当顧客.xlsxを読み込み、キーワードベースで比較分析を行い、
Excelファイルとマークダウンファイル（サマリー）を出力する
"""
import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Hyperlink

import config
from analyzer import coerce_numeric, normalize_columns


def load_data_file(path: Path) -> pd.DataFrame:
    """データファイルを読み込む"""
    if not path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {path}")
    
    if path.suffix.lower() in {".xlsx", ".xls"}:
        df = pd.read_excel(path, engine="openpyxl")
    else:
        df = pd.read_csv(path, encoding="utf-8-sig", thousands=",")
    
    df = normalize_columns(df)
    
    # 必須列のチェック（keywordとcampaign_nameは必須）
    required = {"keyword", "campaign_name"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"{path.name} に必要な列が見つかりません: {', '.join(missing)}")
    
    # 数値列を変換
    numeric_cols = ["avg_position", "imp", "click"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = coerce_numeric(df[col])
    
    df["keyword"] = df["keyword"].astype(str)
    df["campaign_name"] = df["campaign_name"].astype(str)
    
    print(f"[LOAD] {path.name}")
    print(f"  行数: {len(df)}")
    print(f"  列: {df.columns.tolist()}")
    
    return df


def extract_keywords(df: pd.DataFrame) -> set:
    """DataFrameからキーワードを抽出（重複除去）"""
    keywords = set(df["keyword"].dropna().unique())
    # 空文字列を除外
    keywords = {kw for kw in keywords if kw and str(kw).strip()}
    return keywords


def classify_keywords(
    competitor_keywords: set, customer_keywords: set
) -> Tuple[set, set, set]:
    """キーワードを3パターンに分類"""
    both = competitor_keywords & customer_keywords
    competitor_only = competitor_keywords - customer_keywords
    customer_only = customer_keywords - competitor_keywords
    return both, competitor_only, customer_only


def aggregate_keyword_data(df: pd.DataFrame, keyword: str) -> Dict:
    """キーワードのアカウント全体データを集計"""
    keyword_data = df[df["keyword"] == keyword].copy()
    
    if len(keyword_data) == 0:
        return {
            "campaign_count": 0,
            "avg_position": None,
            "imp_total": 0,
            "click_total": 0,
        }
    
    # 平均表示順位は最小値を使用（より良い順位を代表値とする）
    avg_position = keyword_data["avg_position"].min() if "avg_position" in keyword_data.columns else None
    
    imp_total = keyword_data["imp"].sum() if "imp" in keyword_data.columns else 0
    click_total = keyword_data["click"].sum() if "click" in keyword_data.columns else 0
    
    return {
        "campaign_count": len(keyword_data["campaign_name"].unique()),
        "campaigns": keyword_data["campaign_name"].unique().tolist(),
        "avg_position": avg_position,
        "imp_total": imp_total,
        "click_total": click_total,
        "detail_data": keyword_data.to_dict("records"),
    }


def calculate_comparison_metrics(
    competitor_data: Dict, customer_data: Dict
) -> Dict:
    """比較指標を計算"""
    position_diff = None
    imp_diff = None
    click_diff = None
    imp_ratio = None
    click_ratio = None
    
    if competitor_data["avg_position"] is not None and customer_data["avg_position"] is not None:
        position_diff = customer_data["avg_position"] - competitor_data["avg_position"]
    
    imp_diff = customer_data["imp_total"] - competitor_data["imp_total"]
    click_diff = customer_data["click_total"] - competitor_data["click_total"]
    
    if competitor_data["imp_total"] > 0:
        imp_ratio = customer_data["imp_total"] / competitor_data["imp_total"]
    
    if competitor_data["click_total"] > 0:
        click_ratio = customer_data["click_total"] / competitor_data["click_total"]
    
    return {
        "position_diff": position_diff,
        "imp_diff": imp_diff,
        "click_diff": click_diff,
        "imp_ratio": imp_ratio,
        "click_ratio": click_ratio,
    }


def determine_advantage(position_diff: float) -> str:
    """優劣を判定"""
    if position_diff is None:
        return "データなし"
    
    if position_diff <= -2:
        return "優位"
    elif position_diff >= 2:
        return "劣位"
    else:
        return "同等"


def analyze_competitor_comparison(
    competitor_file: Path, customer_file: Path, output_dir: Path = None
):
    """競合比較分析を実行"""
    if output_dir is None:
        output_dir = Path(__file__).parent / "output"
    else:
        output_dir = Path(output_dir)
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # データ読み込み
    print("\n=== データ読み込み ===")
    competitor_df = load_data_file(competitor_file)
    customer_df = load_data_file(customer_file)
    
    # キーワード抽出
    print("\n=== キーワード抽出 ===")
    competitor_keywords = extract_keywords(competitor_df)
    customer_keywords = extract_keywords(customer_df)
    
    print(f"競合のキーワード数: {len(competitor_keywords)}")
    print(f"担当顧客のキーワード数: {len(customer_keywords)}")
    
    # キーワード分類
    both_keywords, competitor_only_keywords, customer_only_keywords = classify_keywords(
        competitor_keywords, customer_keywords
    )
    
    print(f"両方に存在: {len(both_keywords)}")
    print(f"競合のみ: {len(competitor_only_keywords)}")
    print(f"担当顧客のみ: {len(customer_only_keywords)}")
    
    # データ集計と比較
    print("\n=== データ集計と比較 ===")
    
    # アカウント全体レベル（両方に存在するキーワード）
    account_level_results = []
    for keyword in sorted(both_keywords):
        competitor_data = aggregate_keyword_data(competitor_df, keyword)
        customer_data = aggregate_keyword_data(customer_df, keyword)
        metrics = calculate_comparison_metrics(competitor_data, customer_data)
        advantage = determine_advantage(metrics["position_diff"])
        
        account_level_results.append({
            "keyword": keyword,
            "pattern": "両方",
            "competitor_campaign_count": competitor_data["campaign_count"],
            "customer_campaign_count": customer_data["campaign_count"],
            "competitor_avg_position": competitor_data["avg_position"],
            "customer_avg_position": customer_data["avg_position"],
            "position_diff": metrics["position_diff"],
            "competitor_imp_total": competitor_data["imp_total"],
            "customer_imp_total": customer_data["imp_total"],
            "imp_diff": metrics["imp_diff"],
            "competitor_click_total": competitor_data["click_total"],
            "customer_click_total": customer_data["click_total"],
            "click_diff": metrics["click_diff"],
            "advantage": advantage,
        })
    
    # 競合のみのキーワード
    competitor_only_results = []
    for keyword in sorted(competitor_only_keywords):
        competitor_data = aggregate_keyword_data(competitor_df, keyword)
        competitor_only_results.append({
            "keyword": keyword,
            "competitor_campaign_name": ", ".join(competitor_data["campaigns"]),
            "competitor_avg_position": competitor_data["avg_position"],
            "competitor_imp": competitor_data["imp_total"],
            "competitor_click": competitor_data["click_total"],
        })
    
    # 担当顧客のみのキーワード
    customer_only_results = []
    for keyword in sorted(customer_only_keywords):
        customer_data = aggregate_keyword_data(customer_df, keyword)
        customer_only_results.append({
            "keyword": keyword,
            "customer_campaign_name": ", ".join(customer_data["campaigns"]),
            "customer_avg_position": customer_data["avg_position"],
            "customer_imp": customer_data["imp_total"],
            "customer_click": customer_data["click_total"],
        })
    
    # キャンペーン×キーワードレベル（両方に存在するキーワードのみ）
    campaign_keyword_results = []
    for keyword in sorted(both_keywords):
        competitor_keyword_data = competitor_df[competitor_df["keyword"] == keyword]
        customer_keyword_data = customer_df[customer_df["keyword"] == keyword]
        
        # すべてのキャンペーン×キーワードの組み合わせを生成
        for _, comp_row in competitor_keyword_data.iterrows():
            for _, cust_row in customer_keyword_data.iterrows():
                comp_pos = comp_row.get("avg_position", None)
                cust_pos = cust_row.get("avg_position", None)
                position_diff = None
                if comp_pos is not None and cust_pos is not None:
                    position_diff = cust_pos - comp_pos
                
                comp_imp = comp_row.get("imp", 0)
                cust_imp = cust_row.get("imp", 0)
                imp_diff = cust_imp - comp_imp
                
                comp_click = comp_row.get("click", 0)
                cust_click = cust_row.get("click", 0)
                click_diff = cust_click - comp_click
                
                advantage = determine_advantage(position_diff)
                
                campaign_keyword_results.append({
                    "keyword": keyword,
                    "competitor_campaign_name": comp_row["campaign_name"],
                    "customer_campaign_name": cust_row["campaign_name"],
                    "competitor_avg_position": comp_pos,
                    "customer_avg_position": cust_pos,
                    "position_diff": position_diff,
                    "competitor_imp": comp_imp,
                    "customer_imp": cust_imp,
                    "imp_diff": imp_diff,
                    "competitor_click": comp_click,
                    "customer_click": cust_click,
                    "click_diff": click_diff,
                    "advantage": advantage,
                })
    
    # キャンペーン別サマリー
    campaign_summary = []
    customer_campaigns = customer_df["campaign_name"].unique()
    for campaign in sorted(customer_campaigns):
        campaign_customer_data = customer_df[customer_df["campaign_name"] == campaign]
        campaign_keywords = set(campaign_customer_data["keyword"].unique())
        comparable_keywords = campaign_keywords & both_keywords
        
        if len(comparable_keywords) == 0:
            continue
        
        advantage_counts = {"優位": 0, "劣位": 0, "同等": 0}
        position_diffs = []
        imp_ratios = []
        
        for keyword in comparable_keywords:
            competitor_data = aggregate_keyword_data(competitor_df, keyword)
            customer_data = aggregate_keyword_data(customer_df, keyword)
            metrics = calculate_comparison_metrics(competitor_data, customer_data)
            advantage = determine_advantage(metrics["position_diff"])
            
            if advantage in advantage_counts:
                advantage_counts[advantage] += 1
            
            if metrics["position_diff"] is not None:
                position_diffs.append(metrics["position_diff"])
            
            if metrics["imp_ratio"] is not None:
                imp_ratios.append(metrics["imp_ratio"])
        
        total_comparable = len(comparable_keywords)
        advantage_rate = (advantage_counts["優位"] / total_comparable * 100) if total_comparable > 0 else 0
        avg_position_diff = sum(position_diffs) / len(position_diffs) if position_diffs else None
        avg_imp_ratio = sum(imp_ratios) / len(imp_ratios) if imp_ratios else None
        
        campaign_summary.append({
            "campaign_name": campaign,
            "comparable_keyword_count": total_comparable,
            "advantage_count": advantage_counts["優位"],
            "disadvantage_count": advantage_counts["劣位"],
            "advantage_rate": advantage_rate,
            "avg_position_diff": avg_position_diff,
            "avg_imp_ratio": avg_imp_ratio,
        })
    
    # 優劣別集計
    advantage_summary = {}
    for result in account_level_results:
        advantage = result["advantage"]
        advantage_summary[advantage] = advantage_summary.get(advantage, 0) + 1
    
    advantage_summary["競合のみ"] = len(competitor_only_keywords)
    advantage_summary["担当顧客のみ"] = len(customer_only_keywords)
    
    total_keywords = len(both_keywords) + len(competitor_only_keywords) + len(customer_only_keywords)
    
    # Excel出力
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = output_dir / f"competitor_comparison_{timestamp}.xlsx"
    print(f"\n=== Excelファイル出力: {excel_file} ===")
    
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        # シート1: アカウント全体比較（全キーワード）
        all_keywords_df = pd.DataFrame(account_level_results)
        if len(competitor_only_results) > 0:
            comp_only_df = pd.DataFrame(competitor_only_results)
            comp_only_df["pattern"] = "競合のみ"
            comp_only_df["advantage"] = "競合のみ"
            all_keywords_df = pd.concat([all_keywords_df, comp_only_df], ignore_index=True)
        if len(customer_only_results) > 0:
            cust_only_df = pd.DataFrame(customer_only_results)
            cust_only_df["pattern"] = "担当顧客のみ"
            cust_only_df["advantage"] = "担当顧客のみ"
            all_keywords_df = pd.concat([all_keywords_df, cust_only_df], ignore_index=True)
        
        all_keywords_df.to_excel(writer, sheet_name="アカウント全体比較", index=False)
        
        # シート2: キャンペーン×キーワード詳細比較
        if len(campaign_keyword_results) > 0:
            pd.DataFrame(campaign_keyword_results).to_excel(
                writer, sheet_name="キャンペーン×キーワード詳細", index=False
            )
        
        # シート3: 競合のみのキーワード詳細
        if len(competitor_only_results) > 0:
            pd.DataFrame(competitor_only_results).to_excel(
                writer, sheet_name="競合のみ", index=False
            )
        
        # シート4: 担当顧客のみのキーワード詳細
        if len(customer_only_results) > 0:
            pd.DataFrame(customer_only_results).to_excel(
                writer, sheet_name="担当顧客のみ", index=False
            )
        
        # シート5: キャンペーン別サマリー
        if len(campaign_summary) > 0:
            pd.DataFrame(campaign_summary).to_excel(
                writer, sheet_name="キャンペーン別サマリー", index=False
            )
        
        # シート6: 優劣別集計
        advantage_df = pd.DataFrame([
            {"advantage": k, "keyword_count": v, "ratio": (v / total_keywords * 100) if total_keywords > 0 else 0}
            for k, v in advantage_summary.items()
        ])
        advantage_df.to_excel(writer, sheet_name="優劣別集計", index=False)
    
    # 条件付き書式を適用
    apply_conditional_formatting(excel_file)
    
    # ハイパーリンクを追加
    add_hyperlinks(excel_file, account_level_results, competitor_only_results, customer_only_results, campaign_keyword_results)
    
    # マークダウンファイル出力
    md_file = output_dir / f"competitor_comparison_summary_{timestamp}.md"
    print(f"\n=== マークダウンファイル出力: {md_file} ===")
    generate_markdown_summary(
        md_file,
        competitor_file.name,
        customer_file.name,
        len(competitor_keywords),
        len(customer_keywords),
        len(both_keywords),
        len(competitor_only_keywords),
        len(customer_only_keywords),
        account_level_results,
        competitor_only_results,
        customer_only_results,
        campaign_summary,
        advantage_summary,
        total_keywords,
    )
    
    print(f"\n=== 処理完了 ===")
    print(f"Excelファイル: {excel_file}")
    print(f"マークダウンファイル: {md_file}")
    
    return excel_file, md_file


def apply_conditional_formatting(excel_file: Path):
    """条件付き書式を適用"""
    wb = load_workbook(excel_file)
    
    # 色の定義
    colors = {
        "優位": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "劣位": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        "同等": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
        "競合のみ": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        "担当顧客のみ": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
        "データなし": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
    }
    
    # シート1: アカウント全体比較
    if "アカウント全体比較" in wb.sheetnames:
        ws = wb["アカウント全体比較"]
        advantage_col = None
        for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            if col[0].value == "advantage" or col[0].value == "優劣判定":
                advantage_col = idx
                break
        
        if advantage_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                advantage = row[advantage_col - 1].value
                if advantage in colors:
                    fill = colors[advantage]
                    for cell in row:
                        cell.fill = fill
    
    # シート2: キャンペーン×キーワード詳細
    if "キャンペーン×キーワード詳細" in wb.sheetnames:
        ws = wb["キャンペーン×キーワード詳細"]
        advantage_col = None
        for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            if col[0].value == "advantage" or col[0].value == "優劣判定":
                advantage_col = idx
                break
        
        if advantage_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                advantage = row[advantage_col - 1].value
                if advantage in colors:
                    fill = colors[advantage]
                    for cell in row:
                        cell.fill = fill
    
    wb.save(excel_file)


def add_hyperlinks(
    excel_file: Path,
    account_level_results: List[Dict],
    competitor_only_results: List[Dict],
    customer_only_results: List[Dict],
    campaign_keyword_results: List[Dict],
):
    """ハイパーリンクを追加"""
    wb = load_workbook(excel_file)
    
    # シート1からシート2/3/4へのハイパーリンク
    if "アカウント全体比較" in wb.sheetnames:
        ws1 = wb["アカウント全体比較"]
        
        # キーワード列を探す
        keyword_col = None
        for idx, col in enumerate(ws1.iter_cols(min_row=1, max_row=1), start=1):
            if col[0].value == "keyword" or col[0].value == "キーワード":
                keyword_col = idx
                break
        
        if keyword_col:
            # パターン列を探す
            pattern_col = None
            for idx, col in enumerate(ws1.iter_cols(min_row=1, max_row=1), start=1):
                if col[0].value == "pattern" or col[0].value == "存在パターン":
                    pattern_col = idx
                    break
            
            for row_idx, row in enumerate(ws1.iter_rows(min_row=2, max_row=ws1.max_row), start=2):
                keyword = row[keyword_col - 1].value
                pattern = row[pattern_col - 1].value if pattern_col else None
                
                if keyword:
                    # リンク先シートを決定
                    if pattern == "両方":
                        target_sheet = "キャンペーン×キーワード詳細"
                        # 該当キーワードの最初の行を探す
                        target_row = 2
                        if target_sheet in wb.sheetnames:
                            ws_target = wb[target_sheet]
                            for r_idx, r in enumerate(ws_target.iter_rows(min_row=2, max_row=ws_target.max_row), start=2):
                                if r[0].value == keyword:
                                    target_row = r_idx
                                    break
                            link = f"#{target_sheet}!A{target_row}"
                        else:
                            link = None
                    elif pattern == "競合のみ":
                        target_sheet = "競合のみ"
                        target_row = 2
                        if target_sheet in wb.sheetnames:
                            ws_target = wb[target_sheet]
                            for r_idx, r in enumerate(ws_target.iter_rows(min_row=2, max_row=ws_target.max_row), start=2):
                                if r[0].value == keyword:
                                    target_row = r_idx
                                    break
                            link = f"#{target_sheet}!A{target_row}"
                        else:
                            link = None
                    elif pattern == "担当顧客のみ":
                        target_sheet = "担当顧客のみ"
                        target_row = 2
                        if target_sheet in wb.sheetnames:
                            ws_target = wb[target_sheet]
                            for r_idx, r in enumerate(ws_target.iter_rows(min_row=2, max_row=ws_target.max_row), start=2):
                                if r[0].value == keyword:
                                    target_row = r_idx
                                    break
                            link = f"#{target_sheet}!A{target_row}"
                        else:
                            link = None
                    else:
                        link = None
                    
                    if link:
                        cell = row[keyword_col - 1]
                        cell.hyperlink = link
                        cell.style = "Hyperlink"
    
    wb.save(excel_file)


def generate_markdown_summary(
    md_file: Path,
    competitor_file_name: str,
    customer_file_name: str,
    competitor_keyword_count: int,
    customer_keyword_count: int,
    both_count: int,
    competitor_only_count: int,
    customer_only_count: int,
    account_level_results: List[Dict],
    competitor_only_results: List[Dict],
    customer_only_results: List[Dict],
    campaign_summary: List[Dict],
    advantage_summary: Dict,
    total_keywords: int,
):
    """マークダウンファイル（サマリー）を生成"""
    timestamp = datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")
    
    with open(md_file, "w", encoding="utf-8") as f:
        f.write(f"# 競合比較分析サマリー\n\n")
        f.write(f"**分析日時**: {timestamp}\n\n")
        f.write(f"**分析対象**: {competitor_file_name} vs {customer_file_name}\n\n")
        
        f.write("## 全体サマリー\n\n")
        f.write(f"- **競合のキーワード数**: {competitor_keyword_count:,}\n")
        f.write(f"- **担当顧客のキーワード数**: {customer_keyword_count:,}\n")
        f.write(f"- **ユニークキーワード数**: {total_keywords:,}\n")
        f.write(f"- **比較可能なキーワード数（両方に存在）**: {both_count:,}\n")
        f.write(f"- **競合のみのキーワード数**: {competitor_only_count:,}\n")
        f.write(f"- **担当顧客のみのキーワード数**: {customer_only_count:,}\n\n")
        
        f.write("## 優劣別集計\n\n")
        f.write("| 優劣判定 | キーワード数 | 割合（%） |\n")
        f.write("|---------|------------|----------|\n")
        for advantage, count in sorted(advantage_summary.items()):
            ratio = (count / total_keywords * 100) if total_keywords > 0 else 0
            f.write(f"| {advantage} | {count:,} | {ratio:.2f} |\n")
        f.write("\n")
        
        # 主要指標の比較（両方に存在するキーワードのみ）
        if len(account_level_results) > 0:
            f.write("## 主要指標の比較（両方に存在するキーワード）\n\n")
            
            # 平均表示順位
            positions_comp = [r["competitor_avg_position"] for r in account_level_results if r["competitor_avg_position"] is not None]
            positions_cust = [r["customer_avg_position"] for r in account_level_results if r["customer_avg_position"] is not None]
            
            if positions_comp and positions_cust:
                avg_pos_comp = sum(positions_comp) / len(positions_comp)
                avg_pos_cust = sum(positions_cust) / len(positions_cust)
                f.write(f"- **平均表示順位**: 競合 {avg_pos_comp:.2f}位 vs 担当顧客 {avg_pos_cust:.2f}位\n")
            
            # 表示回数合計
            total_imp_comp = sum(r["competitor_imp_total"] for r in account_level_results)
            total_imp_cust = sum(r["customer_imp_total"] for r in account_level_results)
            f.write(f"- **表示回数合計**: 競合 {total_imp_comp:,} vs 担当顧客 {total_imp_cust:,}\n")
            
            # クリック数合計
            total_click_comp = sum(r["competitor_click_total"] for r in account_level_results)
            total_click_cust = sum(r["customer_click_total"] for r in account_level_results)
            f.write(f"- **クリック数合計**: 競合 {total_click_comp:,} vs 担当顧客 {total_click_cust:,}\n\n")
        
        # トップ10リスト
        if len(account_level_results) > 0:
            f.write("## トップ10リスト\n\n")
            
            # 担当顧客が最も優位なキーワード
            advantage_keywords = [r for r in account_level_results if r["advantage"] == "優位"]
            if advantage_keywords:
                advantage_keywords_sorted = sorted(
                    advantage_keywords,
                    key=lambda x: x["position_diff"] if x["position_diff"] is not None else 999
                )[:10]
                f.write("### 担当顧客が最も優位なキーワード（順位差が最も小さい）トップ10\n\n")
                f.write("| キーワード | 競合順位 | 担当顧客順位 | 順位差 |\n")
                f.write("|----------|---------|------------|--------|\n")
                for r in advantage_keywords_sorted:
                    comp_pos = r["competitor_avg_position"] if r["competitor_avg_position"] is not None else "-"
                    cust_pos = r["customer_avg_position"] if r["customer_avg_position"] is not None else "-"
                    diff = r["position_diff"] if r["position_diff"] is not None else "-"
                    f.write(f"| {r['keyword']} | {comp_pos} | {cust_pos} | {diff} |\n")
                f.write("\n")
            
            # 担当顧客が最も劣位なキーワード
            disadvantage_keywords = [r for r in account_level_results if r["advantage"] == "劣位"]
            if disadvantage_keywords:
                disadvantage_keywords_sorted = sorted(
                    disadvantage_keywords,
                    key=lambda x: x["position_diff"] if x["position_diff"] is not None else -999,
                    reverse=True
                )[:10]
                f.write("### 担当顧客が最も劣位なキーワード（順位差が最も大きい）トップ10\n\n")
                f.write("| キーワード | 競合順位 | 担当顧客順位 | 順位差 |\n")
                f.write("|----------|---------|------------|--------|\n")
                for r in disadvantage_keywords_sorted:
                    comp_pos = r["competitor_avg_position"] if r["competitor_avg_position"] is not None else "-"
                    cust_pos = r["customer_avg_position"] if r["customer_avg_position"] is not None else "-"
                    diff = r["position_diff"] if r["position_diff"] is not None else "-"
                    f.write(f"| {r['keyword']} | {comp_pos} | {cust_pos} | {diff} |\n")
                f.write("\n")
        
        # 表示回数・クリック数差分リスト
        if len(account_level_results) > 0:
            # 表示回数差が大きいキーワード（担当顧客が競合より多い）トップ10
            imp_diff_positive = sorted(
                [r for r in account_level_results if r.get("imp_diff", 0) > 0],
                key=lambda x: x.get("imp_diff", 0),
                reverse=True
            )[:10]
            if imp_diff_positive:
                f.write("### 表示回数差が大きいキーワード（担当顧客が競合より多い）トップ10\n\n")
                f.write("| キーワード | 競合表示回数 | 担当顧客表示回数 | 表示回数差 |\n")
                f.write("|----------|------------|--------------|----------|\n")
                for r in imp_diff_positive:
                    comp_imp = r.get("competitor_imp_total", 0)
                    cust_imp = r.get("customer_imp_total", 0)
                    diff = r.get("imp_diff", 0)
                    f.write(f"| {r['keyword']} | {comp_imp:,} | {cust_imp:,} | {diff:+,} |\n")
                f.write("\n")
            
            # 表示回数差が大きいキーワード（競合が担当顧客より多い）トップ10
            imp_diff_negative = sorted(
                [r for r in account_level_results if r.get("imp_diff", 0) < 0],
                key=lambda x: abs(x.get("imp_diff", 0)),
                reverse=True
            )[:10]
            if imp_diff_negative:
                f.write("### 表示回数差が大きいキーワード（競合が担当顧客より多い）トップ10\n\n")
                f.write("| キーワード | 競合表示回数 | 担当顧客表示回数 | 表示回数差 |\n")
                f.write("|----------|------------|--------------|----------|\n")
                for r in imp_diff_negative:
                    comp_imp = r.get("competitor_imp_total", 0)
                    cust_imp = r.get("customer_imp_total", 0)
                    diff = r.get("imp_diff", 0)
                    f.write(f"| {r['keyword']} | {comp_imp:,} | {cust_imp:,} | {diff:,} |\n")
                f.write("\n")
            
            # クリック数差が大きいキーワード（担当顧客が競合より多い）トップ10
            click_diff_positive = sorted(
                [r for r in account_level_results if r.get("click_diff", 0) > 0],
                key=lambda x: x.get("click_diff", 0),
                reverse=True
            )[:10]
            if click_diff_positive:
                f.write("### クリック数差が大きいキーワード（担当顧客が競合より多い）トップ10\n\n")
                f.write("| キーワード | 競合クリック数 | 担当顧客クリック数 | クリック数差 |\n")
                f.write("|----------|------------|--------------|----------|\n")
                for r in click_diff_positive:
                    comp_click = r.get("competitor_click_total", 0)
                    cust_click = r.get("customer_click_total", 0)
                    diff = r.get("click_diff", 0)
                    f.write(f"| {r['keyword']} | {comp_click:,} | {cust_click:,} | {diff:+,} |\n")
                f.write("\n")
            
            # クリック数差が大きいキーワード（競合が担当顧客より多い）トップ10
            click_diff_negative = sorted(
                [r for r in account_level_results if r.get("click_diff", 0) < 0],
                key=lambda x: abs(x.get("click_diff", 0)),
                reverse=True
            )[:10]
            if click_diff_negative:
                f.write("### クリック数差が大きいキーワード（競合が担当顧客より多い）トップ10\n\n")
                f.write("| キーワード | 競合クリック数 | 担当顧客クリック数 | クリック数差 |\n")
                f.write("|----------|------------|--------------|----------|\n")
                for r in click_diff_negative:
                    comp_click = r.get("competitor_click_total", 0)
                    cust_click = r.get("customer_click_total", 0)
                    diff = r.get("click_diff", 0)
                    f.write(f"| {r['keyword']} | {comp_click:,} | {cust_click:,} | {diff:,} |\n")
                f.write("\n")
        
        # 競合のみのキーワード（表示回数が多い順）
        if len(competitor_only_results) > 0:
            competitor_only_sorted = sorted(
                competitor_only_results,
                key=lambda x: x.get("competitor_imp", 0),
                reverse=True
            )[:10]
            f.write("### 競合のみのキーワード（表示回数が多い順）トップ10\n\n")
            f.write("| キーワード | 競合順位 | 表示回数 | クリック数 |\n")
            f.write("|----------|---------|---------|----------|\n")
            for r in competitor_only_sorted:
                pos = r["competitor_avg_position"] if r["competitor_avg_position"] is not None else "-"
                imp = r.get("competitor_imp", 0)
                click = r.get("competitor_click", 0)
                f.write(f"| {r['keyword']} | {pos} | {imp:,} | {click:,} |\n")
            f.write("\n")
        
        # 担当顧客のみのキーワード（表示回数が多い順）
        if len(customer_only_results) > 0:
            customer_only_sorted = sorted(
                customer_only_results,
                key=lambda x: x.get("customer_imp", 0),
                reverse=True
            )[:10]
            f.write("### 担当顧客のみのキーワード（表示回数が多い順）トップ10\n\n")
            f.write("| キーワード | 担当顧客順位 | 表示回数 | クリック数 |\n")
            f.write("|----------|------------|---------|----------|\n")
            for r in customer_only_sorted:
                pos = r["customer_avg_position"] if r["customer_avg_position"] is not None else "-"
                imp = r.get("customer_imp", 0)
                click = r.get("customer_click", 0)
                f.write(f"| {r['keyword']} | {pos} | {imp:,} | {click:,} |\n")
            f.write("\n")
        
        # キャンペーン別サマリー
        if len(campaign_summary) > 0:
            f.write("## キャンペーン別サマリー\n\n")
            f.write("| キャンペーン名 | 比較可能キーワード数 | 優位数 | 劣位数 | 優位率（%） | 平均順位差 |\n")
            f.write("|------------|------------------|-------|-------|-----------|----------|\n")
            for cs in sorted(campaign_summary, key=lambda x: x.get("advantage_rate", 0), reverse=True):
                avg_diff = cs["avg_position_diff"] if cs["avg_position_diff"] is not None else "-"
                f.write(f"| {cs['campaign_name']} | {cs['comparable_keyword_count']} | {cs['advantage_count']} | {cs['disadvantage_count']} | {cs['advantage_rate']:.2f} | {avg_diff} |\n")
            f.write("\n")
        
        # 発見事項
        f.write("## 発見事項\n\n")
        f.write("### 機会損失（競合のみのキーワード）\n\n")
        f.write(f"- 競合が取得しているが担当顧客が取得していないキーワードが **{competitor_only_count}** 件あります\n")
        if competitor_only_count > 0:
            f.write("- これらのキーワードは競合に優位性を与えている可能性があります\n\n")
        
        f.write("### 優位性（担当顧客のみのキーワード）\n\n")
        f.write(f"- 担当顧客が取得しているが競合が取得していないキーワードが **{customer_only_count}** 件あります\n")
        if customer_only_count > 0:
            f.write("- これらのキーワードは担当顧客の優位性を示しています\n\n")
        
        if len(account_level_results) > 0:
            advantage_count = sum(1 for r in account_level_results if r["advantage"] == "優位")
            disadvantage_count = sum(1 for r in account_level_results if r["advantage"] == "劣位")
            f.write("### 比較結果（両方に存在するキーワード）\n\n")
            f.write(f"- 担当顧客が優位なキーワード: **{advantage_count}** 件\n")
            f.write(f"- 担当顧客が劣位なキーワード: **{disadvantage_count}** 件\n")
            f.write(f"- 同等なキーワード: **{len(account_level_results) - advantage_count - disadvantage_count}** 件\n\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="競合比較分析スクリプト")
    parser.add_argument("competitor_file", type=Path, nargs="?", default=None, help="競合データファイル（競合.xlsx）")
    parser.add_argument("customer_file", type=Path, nargs="?", default=None, help="担当顧客データファイル（担当顧客.xlsx）")
    parser.add_argument("--output-dir", type=Path, default=None, help="出力ディレクトリ（デフォルト: output/）")
    
    args = parser.parse_args()
    
    script_dir = Path(__file__).parent
    
    # デフォルトファイルパス
    if args.competitor_file is None:
        args.competitor_file = script_dir / "競合.xlsx"
    if args.customer_file is None:
        args.customer_file = script_dir / "担当顧客.xlsx"
    
    try:
        analyze_competitor_comparison(
            args.competitor_file,
            args.customer_file,
            args.output_dir,
        )
    except Exception as e:
        print(f"エラーが発生しました: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
